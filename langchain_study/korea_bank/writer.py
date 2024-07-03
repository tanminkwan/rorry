import openpyxl
import os

class ExcelWriter:
    def __init__(self, file_name):
        self._file_name = file_name
        self._excel_data = None

    def __enter__(self):
        # 파일이 존재하지 않는 경우 새로운 워크북 생성
        if not os.path.exists(self._file_name):
            self._excel_data = openpyxl.Workbook()
            self._excel_data.active.title = "Sheet1"
        else:
            self._excel_data = openpyxl.load_workbook(self._file_name)
        
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        if self._excel_data:
            self._excel_data.save(self._file_name)  # 변경 사항을 파일에 저장
            self._excel_data.close()

    def set_row(self, **kwargs):
        if not self._excel_data:
            raise Exception("Excel file is not open.")

        sheet = self._excel_data.active

        # 첫 번째 행이 비어 있는 경우
        if sheet.max_row <= 1:
            # kwargs의 key 값을 헤더로 사용
            headers = list(kwargs.keys())
            # 헤더를 첫 번째 행에 추가
            sheet.append(headers)

        # 헤더가 있는 경우 첫 번째 행에서 가져옴
        headers = [cell.value for cell in sheet[1]]

        # 새로운 행 추가
        new_row = []
        for header in headers:
            if header in kwargs:
                new_row.append(kwargs[header])
            else:
                new_row.append(None)

        sheet.append(new_row)
        self._excel_data.save(self._file_name)  # 변경 사항을 파일에 저장

# 사용 예시
if __name__ == "__main__":
    excel_file = "example.xlsx"
    
    with Writer(excel_file) as writer:
        writer.set_row(Name="Dave", Age=28, City="San Francisco")
        writer.set_row(Name="Emma", Age=32, City="Toronto")
